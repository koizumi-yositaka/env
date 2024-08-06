import openpyxl
import os




# ヘッダーの位置
input_start_row = 3
input_start_col = 2
output_start_row = 3
output_start_col = 6


excel_folder="User/input/excel"

def get_excel_data() -> list[object]:
    excel_data=[]
    files=get_all_files_in_directory(excel_folder)
    file_path =f"{excel_folder}/{files[0]}"
    workbook = openpyxl.load_workbook(file_path)
    for sheet_name in workbook.sheetnames:
        if sheet_name!="_temp": 
            sheet=workbook[sheet_name]
            input_data,output_data=get_tabledata_from_sheet(sheet)
            excel_data.append({"input_data":input_data,"output_data":output_data,"sheet_name":sheet_name})
        
    return excel_data

def get_all_files_in_directory(directory_path):
    # ファイル名を格納するリスト
    file_names = []

    # ディレクトリ内のすべてのファイル名を取得
    for root, dirs, files in os.walk(directory_path):
        for file in files:
            if file.endswith('.xlsx') or file.endswith('.xls'):file_names.append(file)
    return file_names

def get_tabledata_from_sheet(sheet):
    
    input_data=[]
    output_data=[]
    current_row = input_start_row + 1
    while True:
        id_cell = sheet.cell(row=current_row, column=input_start_col)
        type_cell = sheet.cell(row=current_row, column=input_start_col + 1)
        value_cell = sheet.cell(row=current_row, column=input_start_col + 2)

        if id_cell.value is None:
            break

        row_data = {"ID":id_cell.value or "", "TYPE":type_cell.value or "", "VALUE":value_cell.value or ""}
        input_data.append(row_data)
        current_row += 1
    
    current_row = output_start_row + 1
    while True:
        id_cell = sheet.cell(row=current_row, column=output_start_col)
        type_cell = sheet.cell(row=current_row, column=output_start_col + 1)
        value_cell = sheet.cell(row=current_row, column=output_start_col + 2)
        expect_cell=sheet.cell(row=current_row, column=output_start_col + 3)

        if id_cell.value is None:
            break

        row_data = {"ID":id_cell.value or "", "TYPE":type_cell.value or "", "VALUE":value_cell.value or "","EXPECT":expect_cell.value or ""}
        output_data.append(row_data)
        current_row += 1
    return input_data,output_data

